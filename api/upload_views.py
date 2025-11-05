from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import status
import openpyxl
from io import BytesIO

from .models import Student, Course, Enrollment, Instructor, Hall, Exam


@api_view(['POST'])
@permission_classes([AllowAny])
@parser_classes([MultiPartParser, FormParser])
def upload_students(request):
    """Upload students from Excel file"""
    try:
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Read Excel file
        wb = openpyxl.load_workbook(BytesIO(file.read()))
        ws = wb.active
        
        created_count = 0
        updated_count = 0
        
        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            
            student_id = row[0]
            name = row[1]
            email = row[2]
            
            student, created = Student.objects.get_or_create(
                studentID=student_id,
                defaults={'name': name, 'email': email}
            )
            
            if created:
                created_count += 1
            else:
                student.name = name
                student.email = email
                student.save()
                updated_count += 1
        
        return Response({
            'message': f'Uploaded successfully. Created: {created_count}, Updated: {updated_count}'
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([AllowAny])
@parser_classes([MultiPartParser, FormParser])
def upload_courses(request):
    """Upload courses from Excel file"""
    try:
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        wb = openpyxl.load_workbook(BytesIO(file.read()))
        ws = wb.active
        
        created_count = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            
            course_name = row[0]
            college = row[1]
            
            Course.objects.get_or_create(
                courseName=course_name,
                college=college
            )
            created_count += 1
        
        return Response({
            'message': f'Uploaded {created_count} courses successfully'
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([AllowAny])
@parser_classes([MultiPartParser, FormParser])
def upload_enrollments(request):
    """Upload student-course enrollments from Excel file"""
    try:
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        wb = openpyxl.load_workbook(BytesIO(file.read()))
        ws = wb.active
        
        created_count = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1]:
                continue
            
            student_id = row[0]
            course_id = row[1]
            
            try:
                student = Student.objects.get(studentID=student_id)
                course = Course.objects.get(courseID=course_id)
                
                Enrollment.objects.get_or_create(
                    student=student,
                    course=course
                )
                created_count += 1
            except (Student.DoesNotExist, Course.DoesNotExist):
                continue
        
        return Response({
            'message': f'Uploaded {created_count} enrollments successfully'
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([AllowAny])
@parser_classes([MultiPartParser, FormParser])
def upload_halls(request):
    """Upload halls from Excel file"""
    try:
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        wb = openpyxl.load_workbook(BytesIO(file.read()))
        ws = wb.active
        
        created_count = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            
            name = row[0]
            capacity = row[1]
            
            Hall.objects.get_or_create(
                name=name,
                defaults={'capacity': capacity}
            )
            created_count += 1
        
        return Response({
            'message': f'Uploaded {created_count} halls successfully'
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
